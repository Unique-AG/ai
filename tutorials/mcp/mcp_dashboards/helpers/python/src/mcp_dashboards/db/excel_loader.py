"""Load Excel workbooks into a per-dataset SQLite database."""

from __future__ import annotations

import logging
import sqlite3
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mcp_dashboards.models import (
    BootstrapSummary,
    BootstrapTableSummary,
    ColumnInfo,
    TableSchema,
    sanitize_identifier,
)
from mcp_dashboards.settings import AppSettings

_log = logging.getLogger(__name__)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _infer_sqlite_type(values: list[Any]) -> str:
    samples = [v for v in values if not _is_blank(v)]
    if not samples:
        return "TEXT"
    if all(isinstance(v, bool) for v in samples):
        return "INTEGER"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in samples):
        return "INTEGER"
    if all(
        isinstance(v, (int, float, Decimal)) and not isinstance(v, bool)
        for v in samples
    ):
        return "REAL"
    if all(isinstance(v, (datetime, date)) for v in samples):
        return "TEXT"

    numeric_ok = True
    all_int = True
    for value in samples:
        if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
            if isinstance(value, float) or (
                isinstance(value, Decimal) and value % 1 != 0
            ):
                all_int = False
            continue
        if isinstance(value, (datetime, date)):
            numeric_ok = False
            break
        try:
            number = Decimal(str(value).strip().replace(",", ""))
            if number % 1 != 0:
                all_int = False
        except (InvalidOperation, ValueError):
            numeric_ok = False
            break
    if numeric_ok:
        return "INTEGER" if all_int else "REAL"
    return "TEXT"


def _normalize_cell(value: Any, sqlite_type: str) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return (
            int(value) if sqlite_type == "INTEGER" and value % 1 == 0 else float(value)
        )
    if sqlite_type == "INTEGER":
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return int(Decimal(str(value).strip().replace(",", "")))
    if sqlite_type == "REAL":
        return (
            float(value)
            if isinstance(value, (int, float))
            else float(Decimal(str(value).strip().replace(",", "")))
        )
    return str(value)


def _unique_column_names(headers: list[str]) -> list[tuple[str, str]]:
    seen: dict[str, int] = {}
    result: list[tuple[str, str]] = []
    for index, header in enumerate(headers):
        raw = str(header).strip() if header is not None else ""
        base = sanitize_identifier(raw, fallback=f"col_{index + 1}")
        count = seen.get(base, 0)
        seen[base] = count + 1
        name = base if count == 0 else f"{base}_{count + 1}"
        result.append((name, raw or name))
    return result


def _non_blank_cells(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    return [cell for cell in row if not _is_blank(cell)]


def _stringish_ratio(row: tuple[Any, ...] | list[Any]) -> float:
    cells = _non_blank_cells(row)
    if not cells:
        return 0.0
    stringish = sum(
        0 if isinstance(cell, bool | int | float | Decimal) else 1 for cell in cells
    )
    return stringish / len(cells)


def find_header_row_index(
    rows: list[tuple[Any, ...]],
    *,
    header_row: int | None = None,
    min_header_cells: int = 3,
    max_scan: int = 40,
) -> int | None:
    """Locate the header row index within an Excel sheet."""
    if not rows:
        return None
    if header_row is not None:
        index = header_row - 1
        if index < 0 or index >= len(rows):
            raise ValueError(
                f"excel_header_row={header_row} is out of range for sheet with {len(rows)} rows"
            )
        if len(_non_blank_cells(rows[index])) == 0:
            raise ValueError(f"excel_header_row={header_row} is blank")
        return index

    best_index: int | None = None
    best_score = -1.0
    for index, row in enumerate(rows[:max_scan]):
        cells = _non_blank_cells(row)
        if len(cells) < min_header_cells:
            continue
        ratio = _stringish_ratio(row)
        if ratio < 0.8:
            continue
        score = float(len(cells)) + ratio
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _sheet_to_schema_and_rows(
    sheet_name: str,
    headers: list[Any],
    data_rows: list[tuple[Any, ...]],
) -> tuple[TableSchema, list[dict[str, Any]]]:
    if not headers or all(_is_blank(header) for header in headers):
        raise ValueError(f"sheet {sheet_name!r} has no header row")
    while headers and _is_blank(headers[-1]):
        headers.pop()

    column_pairs = _unique_column_names(
        ["" if _is_blank(header) else str(header) for header in headers]
    )
    table_name = sanitize_identifier(sheet_name, fallback="sheet")
    width = len(column_pairs)

    aligned: list[list[Any]] = []
    for row in data_rows:
        cells = list(row[:width]) + [None] * max(0, width - len(row))
        if not all(_is_blank(cell) for cell in cells):
            aligned.append(cells)

    columns: list[ColumnInfo] = []
    has_id = any(name == "id" for name, _ in column_pairs)
    for column_index, (name, source_header) in enumerate(column_pairs):
        values = [row[column_index] for row in aligned]
        sqlite_type = _infer_sqlite_type(values)
        is_pk = name == "id"
        columns.append(
            ColumnInfo(
                name=name,
                sqlite_type=sqlite_type if not is_pk else "INTEGER",
                nullable=not is_pk,
                primary_key=is_pk,
                autoincrement=False,
                source_header=source_header,
            )
        )

    if not has_id:
        columns.insert(
            0,
            ColumnInfo(
                name="row_id",
                sqlite_type="INTEGER",
                nullable=False,
                primary_key=True,
                autoincrement=True,
                source_header=None,
            ),
        )

    schema = TableSchema(name=table_name, columns=columns, source_sheet=sheet_name)
    records: list[dict[str, Any]] = []
    for cells in aligned:
        record: dict[str, Any] = {}
        for column_index, (name, _) in enumerate(column_pairs):
            column = schema.get_column(name)
            assert column is not None
            record[name] = _normalize_cell(cells[column_index], column.sqlite_type)
        records.append(record)
    return schema, records


def read_excel_workbook(
    excel_path: Path,
    *,
    header_row: int | None = None,
    min_header_cells: int = 3,
) -> list[tuple[TableSchema, list[dict[str, Any]]]]:
    """Parse every usable sheet into `(schema, rows)` pairs."""
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel workbook not found: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    parsed: list[tuple[TableSchema, list[dict[str, Any]]]] = []
    try:
        for sheet_name in workbook.sheetnames:
            rows = [
                tuple(row) for row in workbook[sheet_name].iter_rows(values_only=True)
            ]
            if not rows:
                _log.warning("Skipping empty sheet %r", sheet_name)
                continue
            try:
                header_index = find_header_row_index(
                    rows,
                    header_row=header_row,
                    min_header_cells=min_header_cells,
                )
            except ValueError as exc:
                _log.warning("Skipping sheet %r: %s", sheet_name, exc)
                continue
            if header_index is None:
                _log.warning(
                    "Skipping sheet %r: no header with >= %d text cells",
                    sheet_name,
                    min_header_cells,
                )
                continue
            schema, records = _sheet_to_schema_and_rows(
                sheet_name, list(rows[header_index]), rows[header_index + 1 :]
            )
            parsed.append((schema, records))
    finally:
        workbook.close()

    if not parsed:
        raise ValueError(f"No usable sheets found in {excel_path}")
    return parsed


def create_table(
    conn: sqlite3.Connection,
    schema: TableSchema,
    *,
    table_constraints: list[str] | None = None,
) -> None:
    """Emit CREATE TABLE for a schema, plus any extra table-level constraints.

    `table_constraints` carries things a `ColumnInfo` cannot express, such as a
    FOREIGN KEY clause for a curated import plan.
    """
    column_defs: list[str] = []
    for column in schema.columns:
        parts = [column.name, column.sqlite_type]
        if column.primary_key:
            parts.append("PRIMARY KEY")
            if column.autoincrement:
                parts.append("AUTOINCREMENT")
        elif not column.nullable:
            parts.append("NOT NULL")
        column_defs.append(" ".join(parts))
    column_defs.extend(table_constraints or [])
    conn.execute(f"CREATE TABLE {schema.name} ({', '.join(column_defs)})")


def insert_rows(
    conn: sqlite3.Connection, schema: TableSchema, rows: list[dict[str, Any]]
) -> int:
    """Insert rows into a table, skipping autoincrement columns."""
    if not rows:
        return 0
    writable = [column.name for column in schema.writable_columns]
    placeholders = ", ".join("?" for _ in writable)
    sql = f"INSERT INTO {schema.name} ({', '.join(writable)}) VALUES ({placeholders})"
    conn.executemany(
        sql, [tuple(row.get(column) for column in writable) for row in rows]
    )
    return len(rows)


def bootstrap_from_excel(
    excel_path: Path,
    db_path: Path,
    *,
    replace: bool = True,
    settings: AppSettings | None = None,
) -> BootstrapSummary:
    """Create or replace one dataset's SQLite DB from its Excel workbook."""
    cfg = settings
    db_path.parent.mkdir(parents=True, exist_ok=True)
    sheets = read_excel_workbook(
        excel_path,
        header_row=cfg.excel_header_row if cfg else None,
        min_header_cells=cfg.excel_min_header_cells if cfg else 3,
    )
    if replace and db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    try:
        summaries: list[BootstrapTableSummary] = []
        with conn:
            for schema, rows in sheets:
                create_table(conn, schema)
                row_count = insert_rows(conn, schema, rows)
                summaries.append(
                    BootstrapTableSummary(
                        table=schema.name,
                        source_sheet=schema.source_sheet,
                        columns=list(schema.columns),
                        row_count=row_count,
                    )
                )
    finally:
        conn.close()

    return BootstrapSummary(excel_path=excel_path, db_path=db_path, tables=summaries)
