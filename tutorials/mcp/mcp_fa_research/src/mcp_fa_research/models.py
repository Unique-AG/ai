"""models.py — the per-name sell-side Excel models, buildable INSIDE the MCP.

Vendored from sandbox python/fa-demo/build_model.py so the nightly / desk-brief
job regenerates the workbooks with fresh data (openpyxl, pure python — no Chrome
needed). Adds the 'Event impact' block on Summary: the overnight event's diff with
PREVIOUS values kept next to NEW (impact.py single source). Path-upsert keeps the
content ids stable. SYNTHETIC — DEMO USE ONLY.
"""

from __future__ import annotations

import io
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import chart_pack
import impact
import note_pack
import seed

# ---- house palette ---------------------------------------------------------------
NAVY = "FF14354A"
GREEN = Font(name="Arial", size=10, color="FF008000")            # hardcoded actuals
BLUE = Font(name="Arial", size=10, color="FF0000FF")             # assumption inputs
BASE = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=13, bold=True, color=NAVY)
DISC = Font(name="Arial", size=8, bold=True, italic=True, color="FFC00000")
HDR = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
YELL = PatternFill("solid", fgColor="FFFFFF00")
GREY = PatternFill("solid", fgColor="FFE8EEF0")
THIN = Border(top=Side(style="thin", color="FFB8C4CC"))

FMT_M = "#,##0;\\(#,##0\\);\\-"
FMT_PCT = "0.0%"
FMT_X = "0.0\\x"

YEARS = ["2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E", "2031E"]
N_HIST = 3            # 2023-25 actuals from chart_pack
N_FCST_RAW = 3        # 2026-28 drivers from chart_pack; 2029-31 fade
COLS = [get_column_letter(2 + i) for i in range(len(YEARS))]   # B..J

# CAPM per name (synthetic but differentiated)
BETA = {"MC FP": 1.00, "KER FP": 1.15, "RMS FP": 0.85, "CFR SW": 0.95,
        "MONC IM": 1.05, "UHR SW": 1.80}   # UHR: the high-beta China call (thesis)
ND_EBITDA_25 = {"MC FP": None, "KER FP": 1.6, "RMS FP": -1.2, "CFR SW": -0.9,
                "MONC IM": -0.7, "UHR SW": 0.3}   # MC uses note_pack net debt
TERM_G = 0.030
CCY_FMT = {"EUR": "\\€#,##0.00", "CHF": '"CHF "#,##0.00'}
CCY_FMT0 = {"EUR": "\\€#,##0", "CHF": '"CHF "#,##0'}


def _num(s):
    """'38,510' → 38510.0 · '(4,880)' → -4880 · '20.0x' → 20.0 · '—' → None."""
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace(",", "").replace("€", "").replace("x", "")
    if s in ("", "—", "-"):
        return None
    if s.startswith("(") and s.endswith(")"):
        return -float(s[1:-1])
    try:
        return float(s)
    except ValueError:
        return None


def _sheet(wb, name, title, sub=""):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws["A2"] = ("SYNTHETIC — DRAFT — NOT INVESTMENT RESEARCH" + (f" · {sub}" if sub else ""))
    ws["A2"].font = DISC
    ws.column_dimensions["A"].width = 34
    for c in COLS:
        ws.column_dimensions[c].width = 11.5
    return ws


def _year_header(ws, row=4, years=YEARS):
    ws.cell(row=row, column=1, value="").fill = NAVY_FILL
    for i, y in enumerate(years):
        c = ws.cell(row=row, column=2 + i, value=y)
        c.font = HDR
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center")


def _label(ws, row, text, bold=False):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD if bold else BASE
    return c


def _put(ws, row, col_i, value, font=BASE, fmt=FMT_M):
    c = ws.cell(row=row, column=2 + col_i, value=value)
    c.font = font
    c.number_format = fmt
    return c


def _build_wb(tk: str):
    cov = next(x for x in seed.COVERAGE if x["ticker"] == tk)
    ov = seed.OVERNIGHT.get(tk) or {}
    raw = chart_pack._RAW[tk]
    ccy = raw["ccy"]
    fmt_ps = CCY_FMT.get(ccy, "#,##0.00")
    fmt_tp = CCY_FMT0.get(ccy, "#,##0")
    tp = ov.get("new_target_price") or cov["target_price"]
    price = cov["price"]
    is_mc = tk == "MC FP"
    pack = note_pack._MC_FP if is_mc else None

    sales_m = [v * 1000 for v in raw["sales"]]          # bn → m
    fcf_m = [v * 1000 for v in raw["fcf"]]
    tax_rate = 0.26
    # calibrate the P&L plugs so the model TIES to the house pack (2025A):
    # netfin = -1% sales; MC: NOSH fixed 497m -> solve minorities; others: minorities
    # 7% -> back-solve NOSH from EPS.
    ebit25 = sales_m[2] * raw["margin"][2] / 100
    pretax25 = ebit25 - 0.01 * sales_m[2]
    if is_mc:
        nosh = 497.0
        minor_pct = round(1 - (raw["eps"][2] * nosh) / (pretax25 * (1 - tax_rate)), 4)
    else:
        minor_pct = 0.07
        nosh = round(pretax25 * (1 - tax_rate) * (1 - minor_pct) / raw["eps"][2], 1)
    payout = round(raw["dps"][2] / raw["eps"][2], 4)
    if is_mc:
        nd25 = 11410.0                                   # note_pack key_financials
    else:
        ebitda25 = sales_m[2] * (raw["margin"][2] / 100 + 0.06)
        nd25 = round(ND_EBITDA_25[tk] * ebitda25, 0)

    # reverse-DCF (house practice): solve the terminal g that reconciles the DCF
    # with the house TP, given the calibrated FCFF path and the name's WACC.
    fade_g = [raw["organic"][3], raw["organic"][4], raw["organic"][5]]
    fade_g += [round(fade_g[-1] * 0.8, 1), round(fade_g[-1] * 0.7, 1), 3.0]
    mfade_c = [raw["margin"][3], raw["margin"][4], raw["margin"][5]]
    mfade_c += [mfade_c[-1]] * 3
    s26 = sales_m[2] * (1 + raw["organic"][3] / 100)
    ebit26 = s26 * raw["margin"][3] / 100
    capex_pct = round((ebit26 + 0.055 * s26 - tax_rate * ebit26
                       - 0.18 * (s26 - sales_m[2]) - fcf_m[3]) / s26, 4)

    def _fcff_path():
        out, s_prev = [], sales_m[2]
        for k in range(6):
            s_now = s_prev * (1 + fade_g[k] / 100)
            eb = s_now * mfade_c[k] / 100
            out.append(eb + 0.055 * s_now - tax_rate * eb - capex_pct * s_now
                       - 0.18 * (s_now - s_prev))
            s_prev = s_now
        return out

    fcff_py = _fcff_path()
    wacc = 0.025 + BETA[tk] * 0.055
    wacc = 0.92 * wacc + 0.08 * 0.028
    best_g, best_gap = TERM_G, float("inf")
    for gi in range(5, 461):
        g = gi / 10000
        if wacc - g < 0.005:
            break
        pv = sum(f / (1 + wacc) ** (k + 1) for k, f in enumerate(fcff_py))
        tvv = fcff_py[-1] * (1 + g) / (wacc - g) / (1 + wacc) ** 6
        v = (pv + tvv - nd25) / nosh
        if abs(v - tp) < best_gap:
            best_gap, best_g = abs(v - tp), g
    term_g = round(best_g, 4)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Assumptions ---------------------------------------------------------------
    a = _sheet(wb, "Assumptions", f"{cov['name']} — Assumptions",
               "all forecast drivers live here")
    _year_header(a)
    r = 5
    if is_mc:
        _label(a, r, "Scenario switch (1 = Base pre-warning · 2 = Post-warning)", bold=True)
        sw = a.cell(row=r, column=2, value=2)
        sw.font = BOLD
        sw.fill = YELL
        r += 1
        _label(a, r, "Organic growth — Base (pre-warning) %")
        base_g = [None, None, None, 4.5, 5.5, 5.5, 4.5, 4.0, 3.5]
        for i, v in enumerate(base_g):
            if v is not None:
                _put(a, r, i, v / 100, BLUE, FMT_PCT)
        base_row = r
        r += 1
    _label(a, r, "Organic growth — " + ("Post-warning %" if is_mc else "sales growth %"))
    grow_row = r
    fade = fade_g
    for i in range(N_HIST):
        _put(a, r, i, raw["organic"][i] / 100, GREEN, FMT_PCT)
    for i, v in enumerate(fade):
        _put(a, r, N_HIST + i, v / 100, BLUE, FMT_PCT)
    r += 1
    if is_mc:
        _label(a, r, "Organic growth — ACTIVE (drives the model)", bold=True)
        for i in range(len(fade)):
            col = COLS[N_HIST + i]
            a[f"{col}{r}"] = f"=IF($B$5=2,{col}{grow_row},{col}{base_row})"
            a[f"{col}{r}"].number_format = FMT_PCT
            a[f"{col}{r}"].font = BOLD
        active_g = r
        r += 1
    else:
        active_g = grow_row
    _label(a, r, "Recurring EBIT margin %")
    marg_row = r
    mfade = mfade_c
    for i in range(N_HIST):
        _put(a, r, i, raw["margin"][i] / 100, GREEN, FMT_PCT)
    for i, v in enumerate(mfade):
        _put(a, r, N_HIST + i, v / 100, BLUE, FMT_PCT)
    r += 2
    for lbl, val in [("Tax rate", tax_rate), ("Payout ratio", payout),
                     ("D&A % of sales", 0.055), ("Capex % of sales", capex_pct),
                     ("Δ WCR % of Δ sales", 0.18),
                     ("Minorities & other (% of post-tax)", minor_pct)]:
        _label(a, r, lbl)
        _put(a, r, 0, val, BLUE, FMT_PCT)
        r += 1
    _label(a, r, "Shares outstanding (m)")
    _put(a, r, 0, nosh, GREEN, "#,##0.0")
    nosh_row = r
    r += 2
    _label(a, r, "BNP CIB Strategy — macro block (house view)", bold=True)
    r += 1
    for lbl, val in [("EURUSD (spot, 23 Jul 26)", "1.14"),
                     ("ECB depo rate FY26e", "1.75%"),
                     ("China GDP FY26e (house)", "+4.6%"),
                     ("China consumption impulse", "stimulus announced — timing the axis "
                      "of the scenario work")]:
        _label(a, r, "   " + lbl)
        a.cell(row=r, column=3, value=val).font = BLUE
        r += 1

    # ---- IncomeStmt ------------------------------------------------------------------
    inc = _sheet(wb, "IncomeStmt", f"{cov['name']} — P&L ({ccy}m)")
    _year_header(inc)
    _label(inc, 5, "Sales", bold=True)
    for i in range(N_HIST):
        _put(inc, 5, i, round(sales_m[i]), GREEN)
    for i in range(N_HIST, len(YEARS)):
        col, prev = COLS[i], COLS[i - 1]
        inc[f"{col}5"] = f"={prev}5*(1+Assumptions!{col}{active_g})"
        inc[f"{col}5"].number_format = FMT_M
    _label(inc, 6, "Recurring EBIT", bold=True)
    for i in range(len(YEARS)):
        col = COLS[i]
        inc[f"{col}6"] = f"={col}5*Assumptions!{col}{marg_row}" if i >= N_HIST else \
            round(sales_m[i] * raw["margin"][i] / 100)
        inc[f"{col}6"].number_format = FMT_M
        if i < N_HIST:
            inc[f"{col}6"].font = GREEN
    _label(inc, 7, "Net financial result")
    _label(inc, 8, "Pre-tax profit")
    _label(inc, 9, f"Tax ({tax_rate:.0%})")
    _label(inc, 10, "Minorities & other")
    _label(inc, 11, "Net profit (adj.)", bold=True)
    _label(inc, 12, "EPS (adj.)", bold=True)
    _label(inc, 13, "DPS")
    for i in range(len(YEARS)):
        col = COLS[i]
        inc[f"{col}7"] = f"=-0.01*{col}5"
        inc[f"{col}8"] = f"={col}6+{col}7"
        inc[f"{col}9"] = f"=-{col}8*Assumptions!$B${marg_row + 2}"
        inc[f"{col}10"] = f"=-({col}8+{col}9)*Assumptions!$B${marg_row + 7}"
        inc[f"{col}11"] = f"={col}8+{col}9+{col}10"
        inc[f"{col}12"] = f"={col}11/Assumptions!$B${nosh_row}"
        inc[f"{col}13"] = f"={col}12*Assumptions!$B${marg_row + 3}"
        for rr in (7, 8, 9, 10, 11):
            inc[f"{col}{rr}"].number_format = FMT_M
        inc[f"{col}12"].number_format = fmt_ps
        inc[f"{col}13"].number_format = fmt_ps
        inc[f"{col}12"].font = BOLD
    _label(inc, 15, "EPS check vs house pack (2025A)")
    inc.cell(row=15, column=2, value=raw["eps"][2]).font = GREEN
    inc["B15"].number_format = fmt_ps
    inc["C15"] = f"=B12-B15"
    inc["C15"].number_format = fmt_ps

    # ---- CashFlow ---------------------------------------------------------------------
    cf = _sheet(wb, "CashFlow", f"{cov['name']} — Cash flow ({ccy}m)")
    _year_header(cf)
    rows = [("EBIT", "=IncomeStmt!{c}6"), ("D&A", "=IncomeStmt!{c}5*Assumptions!$B${da}"),
            ("EBITDA", "={c}5+{c}6"), ("Tax paid", "=-{c}5*Assumptions!$B${tax}"),
            ("Capex", "=-IncomeStmt!{c}5*Assumptions!$B${cap}"),
            ("Δ WCR", "=-(IncomeStmt!{c}5-{p}5_INC)*Assumptions!$B${wcr}"),
            ("Free cash flow", "={c}7+{c}8+{c}9+{c}10")]
    da_row, tax_row, cap_row, wcr_row = marg_row + 4, marg_row + 2, marg_row + 5, marg_row + 6
    for j, (lbl, _) in enumerate(rows):
        _label(cf, 5 + j, lbl, bold=lbl in ("EBITDA", "Free cash flow"))
    for i in range(len(YEARS)):
        col = COLS[i]
        prev = COLS[i - 1] if i else COLS[0]
        cf[f"{col}5"] = f"=IncomeStmt!{col}6"
        cf[f"{col}6"] = f"=IncomeStmt!{col}5*Assumptions!$B${da_row}"
        cf[f"{col}7"] = f"={col}5+{col}6"
        cf[f"{col}8"] = f"=-{col}5*Assumptions!$B${tax_row}"
        cf[f"{col}9"] = f"=-IncomeStmt!{col}5*Assumptions!$B${cap_row}"
        if i == 0:
            cf[f"{col}10"] = 0
        else:
            cf[f"{col}10"] = f"=-(IncomeStmt!{col}5-IncomeStmt!{prev}5)*Assumptions!$B${wcr_row}"
        if i < N_HIST:
            cf[f"{col}11"] = round(fcf_m[i])
            cf[f"{col}11"].font = GREEN
        else:
            cf[f"{col}11"] = f"={col}7+{col}8+{col}9+{col}10"
        for rr in range(5, 12):
            cf[f"{col}{rr}"].number_format = FMT_M

    # ---- WACC ---------------------------------------------------------------------------
    wa = _sheet(wb, "WACC", f"{cov['name']} — Cost of capital")
    vals = [("Risk-free rate", 0.025, FMT_PCT, BLUE), ("Equity risk premium", 0.055, FMT_PCT, BLUE),
            ("Beta", BETA[tk], "0.00", BLUE), ("Cost of equity", None, FMT_PCT, BASE),
            ("Cost of debt (post-tax)", 0.028, FMT_PCT, BLUE), ("Equity weight", 0.92, FMT_PCT, BLUE),
            ("WACC", None, FMT_PCT, BOLD), ("Terminal growth g (house calibration — reverse-DCF to the TP)", term_g, FMT_PCT, BLUE)]
    for j, (lbl, v, fmt, fnt) in enumerate(vals):
        _label(wa, 5 + j, lbl, bold=lbl == "WACC")
        c = wa.cell(row=5 + j, column=2)
        if lbl == "Cost of equity":
            c.value = "=B5+B7*B6"
        elif lbl == "WACC":
            c.value = "=B10*B8+(1-B10)*B9"
        else:
            c.value = v
        c.number_format = fmt
        c.font = fnt

    # ---- DCF ---------------------------------------------------------------------------
    d = _sheet(wb, "DCF", f"{cov['name']} — DCF ({ccy}m)")
    fy = YEARS[N_HIST:]
    _year_header(d, 4, fy)
    fcols = [get_column_letter(2 + i) for i in range(len(fy))]
    _label(d, 5, "FCFF")
    _label(d, 6, "Discount factor")
    _label(d, 7, "PV of FCFF", bold=True)
    for i, col in enumerate(fcols):
        src = COLS[N_HIST + i]
        d[f"{col}5"] = f"=CashFlow!{src}11"
        d[f"{col}6"] = f"=1/(1+WACC!$B$11)^{i + 1}"
        d[f"{col}7"] = f"={col}5*{col}6"
        d[f"{col}5"].number_format = FMT_M
        d[f"{col}6"].number_format = "0.000"
        d[f"{col}7"].number_format = FMT_M
    last = fcols[-1]
    items = [
        ("Phase I — PV of explicit FCFF", f"=SUM(B7:{last}7)", FMT_M),
        ("Terminal value (g on final FCFF)", f"={last}5*(1+WACC!$B$12)/(WACC!$B$11-WACC!$B$12)", FMT_M),
        ("PV of terminal value", f"=B10*{last}6", FMT_M),
        ("Enterprise value", "=B9+B11", FMT_M),
        ("Net (debt) cash 2025A", -nd25, FMT_M),
        ("Equity value", "=B12+B13", FMT_M),
        ("Shares (m)", f"=Assumptions!$B${nosh_row}", "#,##0.0"),
        ("DCF value per share", "=B14/B15", fmt_tp),
    ]
    for j, (lbl, v, fmt) in enumerate(items):
        _label(d, 9 + j, lbl, bold=lbl in ("Enterprise value", "DCF value per share"))
        c = d.cell(row=9 + j, column=2, value=v)
        c.number_format = fmt
        if lbl == "Net (debt) cash 2025A":
            c.font = GREEN
        if lbl == "DCF value per share":
            c.fill = GREY

    # ---- Scenario (sensitivity, computed at build) --------------------------------------
    sc = _sheet(wb, "Scenario", f"{cov['name']} — Value/share sensitivity",
                "computed at build — recompute live via the exane-financial-model skill")
    waccs = [round(wacc + 0.005 * (i - 2), 4) for i in range(5)]
    gs = [round(term_g + 0.005 * (i - 2), 4) for i in range(5)]
    fcff = fcff_py
    sc.cell(row=4, column=1, value="WACC \\ g").font = HDR
    sc.cell(row=4, column=1).fill = NAVY_FILL
    for j, g in enumerate(gs):
        c = sc.cell(row=4, column=2 + j, value=f"{g:.1%}")
        c.font = HDR
        c.fill = NAVY_FILL
    for i, w in enumerate(waccs):
        sc.cell(row=5 + i, column=1, value=f"{w:.1%}").font = BOLD
        for j, g in enumerate(gs):
            if w <= g + 0.005:
                sc.cell(row=5 + i, column=2 + j, value="n.m.")
                continue
            pv = sum(f / (1 + w) ** (k + 1) for k, f in enumerate(fcff))
            tv = fcff[-1] * (1 + g) / (w - g) / (1 + w) ** len(fcff)
            v = (pv + tv - nd25) / nosh
            c = sc.cell(row=5 + i, column=2 + j, value=round(v, 0))
            c.number_format = fmt_tp
    sc.cell(row=12, column=1, value=f"House 12m target price: {tp:,.0f} {ccy} "
            "(triangulated DCF × SOTP × peers — see Summary)").font = BOLD

    # ---- Peers (coverage universe) -------------------------------------------------------
    pe = _sheet(wb, "Peers", "Coverage universe — house view")
    hdr = ["Name", "Ticker", "Rating", "Price", "12m TP", "Upside", "P/E 26E", "P/E 27E"]
    for j, h in enumerate(hdr):
        c = pe.cell(row=4, column=1 + j, value=h)
        c.font = HDR
        c.fill = NAVY_FILL
    for i, c0 in enumerate(seed.COVERAGE):
        ovx = seed.OVERNIGHT.get(c0["ticker"]) or {}
        tpx = ovx.get("new_target_price") or c0["target_price"]
        rr = chart_pack._RAW[c0["ticker"]]
        row = 5 + i
        vals = [c0["name"], c0["ticker"], c0["rating"], c0["price"], tpx, None,
                round(c0["price"] / rr["eps"][3], 1), round(c0["price"] / rr["eps"][4], 1)]
        for j, v in enumerate(vals):
            c = pe.cell(row=row, column=1 + j, value=v)
            c.font = BOLD if c0["ticker"] == tk else BASE
        pe.cell(row=row, column=6, value=f"=E{row}/D{row}-1").number_format = "+0.0%;-0.0%"
        pe.cell(row=row, column=7).number_format = FMT_X
        pe.cell(row=row, column=8).number_format = FMT_X
    pe.column_dimensions["A"].width = 18

    # ---- MC FP extras: Divisional + SOTP + EstVsCons -------------------------------------
    if is_mc and pack:
        seg = pack["segment_details"]
        dv = _sheet(wb, "Divisional", "LVMH — Divisional build-up (EURm)")
        for j, h in enumerate(seg["header"]):
            c = dv.cell(row=4, column=1 + j, value=h)
            c.font = HDR
            c.fill = NAVY_FILL
        r = 5
        for rowvals in seg["rows"]:
            for j, v in enumerate(rowvals):
                n = _num(v) if j else v
                c = dv.cell(row=r, column=1 + j, value=n if j else v)
                if j:
                    c.number_format = FMT_M if abs(n or 0) > 100 else FMT_PCT
                    c.font = GREEN if j <= 2 else BLUE
            r += 1
        dv.cell(row=r, column=1, value="Group (sum of divisions)").font = BOLD
        for j in range(1, len(seg["header"])):
            col = get_column_letter(1 + j)
            dv.cell(row=r, column=1 + j, value=f"=SUM({col}5:{col}{r - 1})").number_format = FMT_M

        so = _sheet(wb, "SOTP", "LVMH — Sum of the parts (2027E EBIT × EV/EBIT)")
        for j, h in enumerate(pack["sotp"]["header"]):
            c = so.cell(row=4, column=1 + j, value=h)
            c.font = HDR
            c.fill = NAVY_FILL
        r = 5
        for rowvals in pack["sotp"]["rows"]:
            for j, v in enumerate(rowvals):
                n = _num(v)
                c = so.cell(row=r, column=1 + j, value=n if (j and n is not None) else v)
                if j and n is not None:
                    c.number_format = FMT_X if "x" in str(v) else FMT_M
                    c.font = BLUE if "x" in str(v) else BASE
            r += 1
        so.column_dimensions["A"].width = 30

        cs = _sheet(wb, "EstVsCons", "LVMH — BNPPE vs consensus")
        for j, h in enumerate(pack["consensus_comparison"]["header"]):
            c = cs.cell(row=4, column=1 + j, value=h)
            c.font = HDR
            c.fill = NAVY_FILL
        r = 5
        for rowvals in pack["consensus_comparison"]["rows"]:
            for j, v in enumerate(rowvals):
                n = _num(v)
                cs.cell(row=r, column=1 + j, value=n if (j and n is not None and "%" not in str(v)) else v)
            r += 1
        cs.column_dimensions["A"].width = 26

    # ---- Summary (first sheet) ------------------------------------------------------------
    su = _sheet(wb, "Summary", f"{cov['name']} ({tk}) — Sell-side model")
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))
    su["A4"] = "Rating"
    su["B4"] = cov["rating"]
    su["B4"].font = BOLD
    su["B4"].fill = GREY
    su["A5"] = "Last price (22 Jul 26)"
    su["B5"] = price
    su["B5"].font = GREEN
    su["B5"].number_format = fmt_tp
    su["A6"] = "House 12m target price"
    su["B6"] = tp
    su["B6"].fill = YELL
    su["B6"].font = BOLD
    su["B6"].number_format = fmt_tp
    su["A7"] = "Upside"
    su["B7"] = "=B6/B5-1"
    su["B7"].number_format = "+0.0%;-0.0%"
    su["A8"] = "DCF value per share (live)"
    su["B8"] = "=DCF!B16"
    su["B8"].number_format = fmt_tp
    su["A9"] = "DCF premium/(discount) to TP"
    su["B9"] = "=B8/B6-1"
    su["B9"].number_format = "+0.0%;-0.0%"
    for rr in range(4, 10):
        su.cell(row=rr, column=1).font = BASE
    su["A11"] = "Key financials"
    su["A11"].font = BOLD
    _year_header(su, 12, YEARS[:6])
    kf = [("Sales ({0}m)".format(ccy), "=IncomeStmt!{c}5", FMT_M),
          ("Recurring EBIT", "=IncomeStmt!{c}6", FMT_M),
          ("EBIT margin", "=IncomeStmt!{c}6/IncomeStmt!{c}5", FMT_PCT),
          ("EPS (adj.)", "=IncomeStmt!{c}12", fmt_ps),
          ("DPS", "=IncomeStmt!{c}13", fmt_ps),
          ("Free cash flow", "=CashFlow!{c}11", FMT_M)]
    for j, (lbl, f, fmt) in enumerate(kf):
        _label(su, 13 + j, lbl)
        for i in range(6):
            col = COLS[i]
            c = su.cell(row=13 + j, column=2 + i, value=f.format(c=col))
            c.number_format = fmt
    su["A20"] = ("Flip Assumptions!B5 (LVMH Base ↔ Post-warning) or any blue driver — "
                 "the whole book re-prices." if is_mc else
                 "Change any blue driver on Assumptions — the whole book re-prices.")
    su["A20"].font = DISC


    # ---- Event impact (previous vs new — the overnight diff) -----------------------
    ev = next((x for x in impact.events() if x["key"] == tk), None)
    if ev and ev["rows"]:
        su["A22"] = f"Event impact — {ev['headline']}"
        su["A22"].font = BOLD
        for j, h in enumerate(("Metric", "Previous", "New", "\u0394")):
            c = su.cell(row=23, column=1 + j, value=h)
            c.font = HDR
            c.fill = NAVY_FILL
        for i2, row in enumerate(ev["rows"]):
            su.cell(row=24 + i2, column=1, value=row["metric"]).font = BASE
            su.cell(row=24 + i2, column=2, value=row["old"]).font = GREEN
            su.cell(row=24 + i2, column=3, value=row["new"]).font = BOLD
            su.cell(row=24 + i2, column=4, value=row["delta"]).font = BASE
        su.cell(row=25 + len(ev["rows"]), column=1, value=ev["status"]).font = DISC

    return wb


def build_model_bytes(tk: str) -> bytes:
    wb = _build_wb(tk)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_all_models() -> dict[str, bytes]:
    """{'names/<TK>/notes/<Name> - Sell-side model (SYNTHETIC).xlsx': bytes} for all 6."""
    out = {}
    for c in seed.COVERAGE:
        fname = f"{c['name']} - Sell-side model (SYNTHETIC).xlsx"
        out[f"names/{c['ticker']}/notes/{fname}"] = build_model_bytes(c["ticker"])
    return out
