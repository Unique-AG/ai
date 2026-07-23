"""Tests for Excel → SQLite bootstrap and schema-driven CRUD."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from mcp_sqlite_excel.db.excel_loader import bootstrap_from_excel
from mcp_sqlite_excel.db.repository import SqliteCrudRepository
from mcp_sqlite_excel.models import FieldMap
from mcp_sqlite_excel.settings import AppSettings


@pytest.mark.ai
def test_AI_bootstrap_creates_tables_from_sheet_names(repo: SqliteCrudRepository) -> None:
    """Purpose: sheets become tables with sanitized names and inferred columns.
    Why this matters: MCP tools must discover schema from Excel without hardcoding columns.
    Setup summary: sample workbook with positions + instruments sheets.
    """
    tables = repo.list_tables()
    assert "positions" in tables
    assert "instruments" in tables

    positions = repo.get_table_schema("positions")
    column_names = {c.name for c in positions.columns}
    assert "row_id" in column_names
    assert "ticker" in column_names
    assert "target_weight" in column_names
    assert positions.pk_column.name == "row_id"
    assert positions.pk_column.autoincrement
    assert positions.primary_key == "row_id"


@pytest.mark.ai
def test_AI_bootstrap_uses_existing_id_as_primary_key(tmp_path: Path) -> None:
    """Purpose: an Excel id column becomes the SQLite primary key.
    Why this matters: workbooks with an id column must not get a synthetic row_id.
    Setup summary: minimal workbook with an id header.
    """
    excel = tmp_path / "with_id.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "items"
    sheet.append(["id", "Name", "Qty"])
    sheet.append([1, "Widget", 3])
    sheet.append([2, "Gadget", 5])
    wb.save(excel)

    settings = AppSettings(excel_path=excel, sqlite_path=tmp_path / "with_id.db", auth_disabled=True)
    summary = bootstrap_from_excel(settings=settings, replace=True)
    assert summary.tables[0].table == "items"

    repo = SqliteCrudRepository(settings=settings)
    schema = repo.get_table_schema("items")
    assert schema.pk_column.name == "id"
    assert "row_id" not in schema.column_names
    assert repo.list_rows("items").total_matching == 2


@pytest.mark.ai
def test_AI_crud_create_update_delete_roundtrip(repo: SqliteCrudRepository) -> None:
    """Purpose: create / update / delete work against Excel-derived columns.
    Why this matters: MCP tools must mutate rows using only discovered schema.
    Setup summary: seeded positions table from the sample workbook.
    """
    created = repo.create_row(
        "positions",
        FieldMap(
            {
                "sleeve": "Equity Long",
                "ticker": "NVDA",
                "instrument": "NVIDIA Corp",
                "direction": "Long",
                "target_weight": 0.05,
                "position_mm": 100,
                "email": "alice@alphabet.example",
            }
        ),
    )
    row_id = created.row["row_id"]

    fetched = repo.get_row("positions", row_id)
    assert fetched.row["ticker"] == "NVDA"

    updated = repo.update_row(
        "positions",
        row_id,
        FieldMap({"direction": "Short", "position_mm": -100}),
    )
    assert updated.row["direction"] == "Short"
    assert updated.row["position_mm"] == -100

    listed = repo.list_rows("positions", filters=FieldMap({"ticker": "NVDA"}))
    assert listed.total_matching == 1

    deleted = repo.delete_row("positions", row_id)
    assert deleted.deleted["ticker"] == "NVDA"
    with pytest.raises(KeyError):
        repo.get_row("positions", row_id)


@pytest.mark.ai
def test_AI_list_rows_rejects_unknown_filter_columns(repo: SqliteCrudRepository) -> None:
    """Purpose: filter keys must match schema columns (SQL injection surface).
    Why this matters: LLM-supplied filter keys must not be interpolated unless known.
    Setup summary: seeded repository.
    """
    with pytest.raises(ValueError, match="Unknown column"):
        repo.list_rows("positions", filters={"drop_table": "1"})


@pytest.mark.ai
def test_AI_reset_from_excel_restores_seed(repo: SqliteCrudRepository) -> None:
    """Purpose: reset rebuilds the DB from Excel and drops demo mutations.
    Why this matters: demos need a one-click restore to a known baseline.
    Setup summary: create an extra row, then reset.
    """
    before = repo.list_rows("positions", limit=2000).total_matching
    repo.create_row(
        "positions",
        {"sleeve": "Rates", "ticker": "ZZZ", "instrument": "Temp", "direction": "Long"},
    )
    assert repo.list_rows("positions", limit=2000).total_matching == before + 1

    summary = repo.reset_from_excel()
    assert any(t.table == "positions" for t in summary.tables)
    assert repo.list_rows("positions", limit=2000).total_matching == before


@pytest.mark.ai
def test_AI_field_map_from_mcp_arg_accepts_dict_and_string() -> None:
    """Purpose: MCP args may arrive as JSON strings or already-parsed dicts.
    Why this matters: tool handlers should accept both shapes from clients.
    Setup summary: none.
    """
    assert FieldMap.from_mcp_arg({"a": 1}, field_name="fields").root == {"a": 1}
    assert FieldMap.from_mcp_arg('{"a": 1}', field_name="fields").root == {"a": 1}
    assert FieldMap.from_mcp_arg(None, field_name="filters").root == {}
    with pytest.raises(ValueError):
        FieldMap.from_mcp_arg("[1,2]", field_name="fields")


@pytest.mark.ai
def test_AI_app_settings_loads_paths_from_env(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """Purpose: AppSettings reads EXCEL_PATH / SQLITE_PATH from the environment.
    Why this matters: demos configure the workbook without code changes.
    Setup summary: monkeypatched env vars pointing at temp paths.
    """
    excel = tmp_path / "custom.xlsx"
    db = tmp_path / "custom.db"
    monkeypatch.setenv("EXCEL_PATH", str(excel))
    monkeypatch.setenv("SQLITE_PATH", str(db))
    monkeypatch.setenv("AUTH_DISABLED", "true")

    cfg = AppSettings()
    assert cfg.excel_path == excel
    assert cfg.sqlite_path == db
    assert cfg.auth_disabled is True


@pytest.mark.ai
def test_AI_bootstrap_skips_title_rows_and_finds_header(tmp_path: Path) -> None:
    """Purpose: preamble title rows are skipped so the real header is used.
    Why this matters: report-style workbooks (e.g. account review) put headers below titles.
    Setup summary: sheet with two title rows, then a wide header + data.
    """
    excel = tmp_path / "preamble.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Clients"
    sheet.append(["Account Review — RM Client Book"])
    sheet.append(["RM: demo subtitle"])
    sheet.append([])
    sheet.append(["Client Ref", "Client Name", "Status", "Portfolio Value"])
    sheet.append(["CH-priv-0187", "Alexander Nesterov", "Escalated", 18400000])
    sheet.append(["CH-priv-0204", "Natalia Morozova", "Escalated", 24750000])
    wb.save(excel)

    settings = AppSettings(excel_path=excel, sqlite_path=tmp_path / "preamble.db", auth_disabled=True)
    summary = bootstrap_from_excel(settings=settings, replace=True)
    assert summary.tables[0].table == "clients"
    cols = {c.name for c in summary.tables[0].columns}
    assert "client_ref" in cols
    assert "client_name" in cols
    assert "account_review_rm_client_book" not in cols

    repo = SqliteCrudRepository(settings=settings)
    listed = repo.list_rows("clients")
    assert listed.total_matching == 2
    assert listed.rows[0]["client_ref"] == "CH-priv-0187"


@pytest.mark.ai
def test_AI_bootstrap_honors_explicit_excel_header_row(tmp_path: Path) -> None:
    """Purpose: EXCEL_HEADER_ROW forces a fixed 1-based header shift.
    Why this matters: auto-detect can be overridden for unusual layouts.
    Setup summary: force header on Excel row 3.
    """
    excel = tmp_path / "forced.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Items"
    sheet.append(["ignore me"])
    sheet.append(["also ignore", "x", "y", "z"])
    sheet.append(["Name", "Qty", "Color"])
    sheet.append(["Widget", 3, "red"])
    wb.save(excel)

    settings = AppSettings(
        excel_path=excel,
        sqlite_path=tmp_path / "forced.db",
        auth_disabled=True,
        excel_header_row=3,
    )
    summary = bootstrap_from_excel(settings=settings, replace=True)
    cols = {c.name for c in summary.tables[0].columns}
    assert cols == {"row_id", "name", "qty", "color"}
    assert summary.tables[0].row_count == 1


@pytest.mark.ai
def test_AI_bootstrap_account_review_dataset_if_present() -> None:
    """Purpose: the account_review_dataset workbook loads Clients + Smart Actions.
    Why this matters: the demo dataset is a report-style Excel with title rows.
    Setup summary: skip if the file is not checked in locally.
    """
    excel = Path(__file__).resolve().parents[1] / "data" / "account_review_dataset.xlsx"
    if not excel.is_file():
        pytest.skip("account_review_dataset.xlsx not present")

    settings = AppSettings(
        excel_path=excel,
        sqlite_path=excel.with_name("account_review_test.db"),
        auth_disabled=True,
    )
    summary = bootstrap_from_excel(settings=settings, replace=True)
    tables = {t.table: t for t in summary.tables}
    assert "clients" in tables
    assert "smart_actions" in tables
    client_cols = {c.name for c in tables["clients"].columns}
    assert "client_ref" in client_cols
    assert "portfolio_value" in client_cols
    assert tables["clients"].row_count == 12
    assert tables["smart_actions"].row_count == 8
    # KPI / Legend are key-value sheets — skipped by min header width
    assert "kpi_summary" not in tables
    assert "legend" not in tables
    settings.sqlite_path.unlink(missing_ok=True)


@pytest.mark.ai
def test_AI_bootstrap_summary_is_pydantic_model(app_settings: AppSettings) -> None:
    """Purpose: bootstrap returns a typed BootstrapSummary model.
    Why this matters: callers should serialize/validate seed results consistently.
    Setup summary: seed from the sample workbook settings fixture.
    """
    summary = bootstrap_from_excel(settings=app_settings, replace=True)
    assert summary.excel_path == app_settings.excel_path
    assert summary.db_path == app_settings.sqlite_path
    dumped = summary.model_dump(mode="json")
    assert "tables" in dumped
    with pytest.raises(ValidationError):
        summary.model_validate({"excel_path": "/x.xlsx", "db_path": "/x.db", "tables": "nope"})
