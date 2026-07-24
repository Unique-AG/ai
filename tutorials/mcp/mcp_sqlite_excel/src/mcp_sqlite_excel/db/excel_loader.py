"""Load Excel workbooks into SQLite with schema inferred from headers and values."""

from __future__ import annotations

import logging
import sqlite3
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mcp_sqlite_excel.models import (
    BootstrapSummary,
    BootstrapTableSummary,
    ColumnInfo,
    TableSchema,
    sanitize_identifier,
)
from mcp_sqlite_excel.settings import AppSettings, get_settings

_log = logging.getLogger(__name__)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _infer_sqlite_type(values: list[Any]) -> str:
    """Infer a SQLite affinity from non-empty sample values."""
    samples = [v for v in values if not _is_blank(v)]
    if not samples:
        return "TEXT"

    if all(isinstance(v, bool) for v in samples):
        return "INTEGER"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in samples):
        return "INTEGER"
    if all(isinstance(v, (int, float, Decimal)) and not isinstance(v, bool) for v in samples):
        return "REAL"
    if all(isinstance(v, (datetime, date)) for v in samples):
        return "TEXT"

    numeric_ok = True
    all_int = True
    for v in samples:
        if isinstance(v, (int, float, Decimal)) and not isinstance(v, bool):
            if isinstance(v, float) or (isinstance(v, Decimal) and v % 1 != 0):
                all_int = False
            continue
        if isinstance(v, (datetime, date)):
            numeric_ok = False
            break
        try:
            d = Decimal(str(v).strip().replace(",", ""))
            if d % 1 != 0:
                all_int = False
        except (InvalidOperation, ValueError):
            numeric_ok = False
            break
    if numeric_ok:
        return "INTEGER" if all_int else "REAL"
    return "TEXT"


def _normalize_cell(value: Any, sqlite_type: str) -> Any:
    """Coerce an Excel cell value into something SQLite can store cleanly."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return int(value) if sqlite_type == "INTEGER" and value % 1 == 0 else float(value)
    if sqlite_type == "INTEGER":
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return int(Decimal(str(value).strip().replace(",", "")))
    if sqlite_type == "REAL":
        return float(value) if isinstance(value, (int, float)) else float(Decimal(str(value).strip().replace(",", "")))
    return str(value)


def _unique_column_names(headers: list[str]) -> list[tuple[str, str]]:
    """Return (sanitized_name, original_header) pairs, uniquified with suffixes."""
    seen: dict[str, int] = {}
    result: list[tuple[str, str]] = []
    for idx, header in enumerate(headers):
        raw = str(header).strip() if header is not None else ""
        base = sanitize_identifier(raw, fallback=f"col_{idx + 1}")
        count = seen.get(base, 0)
        seen[base] = count + 1
        name = base if count == 0 else f"{base}_{count + 1}"
        result.append((name, raw or name))
    return result


def _non_blank_cells(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    return [c for c in row if not _is_blank(c)]


def _stringish_ratio(row: tuple[Any, ...] | list[Any]) -> float:
    cells = _non_blank_cells(row)
    if not cells:
        return 0.0
    stringish = 0
    for cell in cells:
        if isinstance(cell, bool):
            continue
        if isinstance(cell, (int, float, Decimal)):
            continue
        stringish += 1
    return stringish / len(cells)


def find_header_row_index(
    rows: list[tuple[Any, ...]],
    *,
    header_row: int | None = None,
    min_header_cells: int = 3,
    max_scan: int = 40,
) -> int | None:
    """Locate the header row index within ``rows``.

    Args:
        rows: All sheet rows (values only).
        header_row: Optional 1-based Excel row number to force (title shift).
        min_header_cells: Minimum non-empty cells for auto-detection.
        max_scan: How many leading rows to consider when auto-detecting.

    Returns:
        0-based index into ``rows``, or ``None`` if no suitable header exists.
    """
    if not rows:
        return None

    if header_row is not None:
        idx = header_row - 1
        if idx < 0 or idx >= len(rows):
            raise ValueError(f"excel_header_row={header_row} is out of range for sheet with {len(rows)} rows")
        if len(_non_blank_cells(rows[idx])) == 0:
            raise ValueError(f"excel_header_row={header_row} is blank")
        return idx

    best_idx: int | None = None
    best_score = -1.0
    for i, row in enumerate(rows[:max_scan]):
        cells = _non_blank_cells(row)
        n = len(cells)
        if n < min_header_cells:
            continue
        ratio = _stringish_ratio(row)
        # Title rows are usually 1 cell; data rows mix numbers. Prefer wide text rows.
        if ratio < 0.8:
            continue
        score = float(n) + ratio
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _sheet_to_schema_and_rows(
    sheet_name: str,
    headers: list[Any],
    data_rows: list[tuple[Any, ...]],
) -> tuple[TableSchema, list[dict[str, Any]]]:
    if not headers or all(_is_blank(h) for h in headers):
        raise ValueError(f"sheet {sheet_name!r} has no header row")

    while headers and _is_blank(headers[-1]):
        headers.pop()

    col_pairs = _unique_column_names(["" if _is_blank(h) else str(h) for h in headers])
    table_name = sanitize_identifier(sheet_name, fallback="sheet")

    width = len(col_pairs)
    aligned: list[list[Any]] = []
    for row in data_rows:
        cells = list(row[:width]) + [None] * max(0, width - len(row))
        if all(_is_blank(c) for c in cells):
            continue
        aligned.append(cells)

    columns: list[ColumnInfo] = []
    has_id = any(name == "id" for name, _ in col_pairs)
    for col_idx, (name, source_header) in enumerate(col_pairs):
        values = [row[col_idx] for row in aligned]
        sqlite_type = _infer_sqlite_type(values)
        is_pk = name == "id"
        columns.append(
            ColumnInfo(
                name=name,
                sqlite_type=sqlite_type if not is_pk else "INTEGER",
                nullable=not is_pk,
                primary_key=is_pk,
                autoincrement=False,
                source_header=source_header,
            )
        )

    if not has_id:
        columns.insert(
            0,
            ColumnInfo(
                name="row_id",
                sqlite_type="INTEGER",
                nullable=False,
                primary_key=True,
                autoincrement=True,
                source_header=None,
            ),
        )

    schema = TableSchema(name=table_name, columns=columns, source_sheet=sheet_name)

    records: list[dict[str, Any]] = []
    for cells in aligned:
        record: dict[str, Any] = {}
        for col_idx, (name, _) in enumerate(col_pairs):
            col = schema.get_column(name)
            assert col is not None
            record[name] = _normalize_cell(cells[col_idx], col.sqlite_type)
        records.append(record)

    return schema, records


def read_excel_workbook(
    excel_path: Path,
    *,
    header_row: int | None = None,
    min_header_cells: int = 3,
) -> list[tuple[TableSchema, list[dict[str, Any]]]]:
    """Parse every usable sheet into (schema, rows).

    Skips title/blank preamble by auto-detecting the header row (or using
    ``header_row`` when provided). Sheets without a wide enough header are skipped.
    """
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel workbook not found: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    parsed: list[tuple[TableSchema, list[dict[str, Any]]]] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
            if not rows:
                _log.warning("Skipping empty sheet %r", sheet_name)
                continue

            try:
                header_idx = find_header_row_index(
                    rows,
                    header_row=header_row,
                    min_header_cells=min_header_cells,
                )
            except ValueError as exc:
                _log.warning("Skipping sheet %r: %s", sheet_name, exc)
                continue

            if header_idx is None:
                _log.warning(
                    "Skipping sheet %r: no header with >= %d text cells",
                    sheet_name,
                    min_header_cells,
                )
                continue

            header_row_values = list(rows[header_idx])
            data_rows = rows[header_idx + 1 :]
            if all(_is_blank(h) for h in header_row_values) and not data_rows:
                _log.warning("Skipping blank sheet %r", sheet_name)
                continue

            _log.info(
                "Sheet %r: using Excel row %d as header (%d non-empty cells)",
                sheet_name,
                header_idx + 1,
                len(_non_blank_cells(header_row_values)),
            )
            schema, records = _sheet_to_schema_and_rows(sheet_name, header_row_values, data_rows)
            parsed.append((schema, records))
    finally:
        workbook.close()

    if not parsed:
        raise ValueError(f"No usable sheets found in {excel_path}")
    return parsed


def _create_table(conn: sqlite3.Connection, schema: TableSchema) -> None:
    col_defs: list[str] = []
    for col in schema.columns:
        parts = [col.name, col.sqlite_type]
        if col.primary_key:
            parts.append("PRIMARY KEY")
            if col.autoincrement:
                parts.append("AUTOINCREMENT")
        elif not col.nullable:
            parts.append("NOT NULL")
        col_defs.append(" ".join(parts))
    ddl = f"CREATE TABLE {schema.name} ({', '.join(col_defs)})"
    conn.execute(ddl)


def _insert_rows(conn: sqlite3.Connection, schema: TableSchema, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    writable = [c.name for c in schema.writable_columns]
    placeholders = ", ".join("?" for _ in writable)
    cols_sql = ", ".join(writable)
    sql = f"INSERT INTO {schema.name} ({cols_sql}) VALUES ({placeholders})"
    values = [tuple(row.get(c) for c in writable) for row in rows]
    conn.executemany(sql, values)
    return len(values)


def bootstrap_from_excel(
    excel_path: Path | None = None,
    db_path: Path | None = None,
    *,
    replace: bool = True,
    settings: AppSettings | None = None,
) -> BootstrapSummary:
    """Create (or replace) the SQLite DB from an Excel workbook."""
    cfg = settings or get_settings()
    excel_path = Path(excel_path or cfg.excel_path)
    db_path = Path(db_path or cfg.sqlite_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    _log.info(
        "Bootstrapping SQLite from Excel: excel_path=%s -> db_path=%s replace=%s",
        excel_path.resolve(),
        db_path.resolve(),
        replace,
    )

    sheets = read_excel_workbook(
        excel_path,
        header_row=cfg.excel_header_row,
        min_header_cells=cfg.excel_min_header_cells,
    )

    if replace and db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    try:
        summary_tables: list[BootstrapTableSummary] = []
        with conn:
            for schema, rows in sheets:
                _create_table(conn, schema)
                count = _insert_rows(conn, schema, rows)
                summary_tables.append(
                    BootstrapTableSummary(
                        table=schema.name,
                        source_sheet=schema.source_sheet,
                        columns=list(schema.columns),
                        row_count=count,
                    )
                )
                _log.info(
                    "Bootstrapped table %s from sheet %r (%d rows, %d columns)",
                    schema.name,
                    schema.source_sheet,
                    count,
                    len(schema.columns),
                )
    finally:
        conn.close()

    return BootstrapSummary(
        excel_path=excel_path,
        db_path=db_path,
        tables=summary_tables,
    )
