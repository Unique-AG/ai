#!/usr/bin/env python3
"""Add columns for four new client-profile sections to account_review_dataset.xlsx.

Why this exists
----------------
The dashboard's client detail page is growing four new *always-visible*
sections (Portfolio and Mandate, Portfolio Performance and Benchmark,
Holdings and Categorization, Suitability Profile) — see
``../data/acount_review/dashboard-v005-astro/src/pages/index.astro``. Unlike the
existing case-specific ``Fig1..3``/``Perf1..3`` columns (populated only for
the one row whose ``rule_code`` needs that exact figure), these four
sections must show real, plausible data for *every* client regardless of
which remediation case (if any) is currently open on their row.

Column design
--------------
- ``Mandate Type`` + ``Mand1..3 Label/Value/Status/Pct`` — a dedicated
  3-row bar figure (Equity/Fixed income/Cash), same shape as the existing
  ``Fig1..3`` columns but under its own prefix so it doesn't collide with
  the case-specific figure a row's ``rule_code`` might still need (e.g.
  doc-expiry's "Documents & KYC", adverse-media's "Screening match").
  Purushottam R Sharma (``R-SUIT-ALLOC``, the live allocation-breach case)
  keeps the exact breach numbers that used to live in ``Fig1..3`` — they
  just move to ``Mand1..3`` since that section is now the generic, always
  shown one.
- ``Perf1..3 Label/Value/Status/Pct`` already exist (added for the
  suit-alloc figure2) and are simply populated for every row here.
- ``Client Categorization`` / ``Category Required`` / ``Category Review
  Status`` + ``Hold1..3 Label/Value/Status/Pct`` — categorization fields
  plus a top-3-holdings bar figure. Aaron Lim Wei Jian's
  (``R-REG-NONDOM``) existing categorization-mismatch story (previously in
  ``Fig1..3``) moves here.
- ``Risk Tolerance`` / ``Investment Horizon`` / ``Knowledge Experience`` /
  ``Last Suitability Test`` / ``Suitability Outcome`` — Markus Brunner's
  (``R-SUIT-REVIEW``) existing snapshot (previously in ``Fig1..3``) moves
  here.

After this script retires a case's use of ``Fig1..3`` (suit-alloc,
reg-change, suit-review), ``cases.json`` no longer declares a
``figure_title``/``figure2_title`` for that case — the content lives in
the generic sections below instead. doc-expiry, adverse-media and
sow-refresh are untouched and keep their own ``Fig1..3`` case figure.

Usage
-----
    uv run python scripts/add_profile_sections.py

Edits ``data/account_review_dataset.xlsx`` in place (only the ``Clients``
sheet). Re-run ``uv run python -m mcp_sqlite_excel.db.excel_loader`` (or
just restart the MCP server / hit ``reset_from_excel``) afterwards to
reload SQLite from the updated workbook.
"""

from __future__ import annotations

import copy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "data" / "account_review_dataset.xlsx"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# fmt: off
NEW_COLUMNS = [
    "Mandate Type",
    "Mand1 Label", "Mand1 Value", "Mand1 Status", "Mand1 Pct",
    "Mand2 Label", "Mand2 Value", "Mand2 Status", "Mand2 Pct",
    "Mand3 Label", "Mand3 Value", "Mand3 Status", "Mand3 Pct",
    "Client Categorization", "Category Required", "Category Review Status",
    "Hold1 Label", "Hold1 Value", "Hold1 Status", "Hold1 Pct",
    "Hold2 Label", "Hold2 Value", "Hold2 Status", "Hold2 Pct",
    "Hold3 Label", "Hold3 Value", "Hold3 Status", "Hold3 Pct",
    "Risk Tolerance", "Investment Horizon", "Knowledge Experience",
    "Last Suitability Test", "Suitability Outcome",
]
# fmt: on

# Keyed by "Client Ref" (stable across re-runs, unlike row number).
ROWS: dict[str, dict[str, object]] = {
    "CH-priv-0187": {  # Alexander Nesterov — Growth, adverse-media (escalated)
        "Mandate Type": "Discretionary — Growth",
        "Mand1 Label": "Equity",
        "Mand1 Value": "66% (ceiling 75%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 66,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "22% (band ≤30%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 22,
        "Mand3 Label": "Cash",
        "Mand3 Value": "12% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 12,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+9.2%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — MSCI ACWI",
        "Perf2 Value": "+8.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "+0.8 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Professional client (per-se)",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — Global Large Cap",
        "Hold1 Value": "38%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 38,
        "Hold2 Label": "Private Equity / Alternatives",
        "Hold2 Value": "24%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 24,
        "Hold3 Label": "Fixed Income — EM Corporate",
        "Hold3 Value": "18%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 18,
        "Risk Tolerance": "Growth",
        "Investment Horizon": "10+ years",
        "Knowledge Experience": "Advanced — active trader",
        "Last Suitability Test": "2026-03-02",
        "Suitability Outcome": "Aligned with Growth mandate",
    },
    "CH-priv-0204": {  # Natalia Morozova — Balanced, PEP (escalated)
        "Mandate Type": "Discretionary — Balanced",
        "Mand1 Label": "Equity",
        "Mand1 Value": "48% (ceiling 60%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 48,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "38% (band ≤45%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 38,
        "Mand3 Label": "Cash",
        "Mand3 Value": "14% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 14,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+5.6%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 60/40 blended",
        "Perf2 Value": "+6.1%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "-0.5 pp",
        "Perf3 Status": "warn",
        "Perf3 Pct": None,
        "Client Categorization": "Retail — MiFID standard",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Fixed Income — UK Gilts",
        "Hold1 Value": "30%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 30,
        "Hold2 Label": "Equities — UK Large Cap",
        "Hold2 Value": "26%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 26,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "14%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 14,
        "Risk Tolerance": "Balanced",
        "Investment Horizon": "5–10 years",
        "Knowledge Experience": "Intermediate",
        "Last Suitability Test": "2025-11-14",
        "Suitability Outcome": "Aligned with Balanced mandate",
    },
    "CH-priv-0231": {  # Dmitry Volkov — Balanced, new adverse-media hit (needs remediation)
        "Mandate Type": "Advisory — Balanced",
        "Mand1 Label": "Equity",
        "Mand1 Value": "55% (ceiling 60%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 55,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "30% (band ≤45%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 30,
        "Mand3 Label": "Cash",
        "Mand3 Value": "15% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 15,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+6.4%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 60/40 blended",
        "Perf2 Value": "+6.1%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "+0.3 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Retail — MiFID standard",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — Swiss Large Cap",
        "Hold1 Value": "32%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 32,
        "Hold2 Label": "Fixed Income — CHF Corporate",
        "Hold2 Value": "28%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 28,
        "Hold3 Label": "Real Estate Funds",
        "Hold3 Value": "12%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 12,
        "Risk Tolerance": "Balanced",
        "Investment Horizon": "5–10 years",
        "Knowledge Experience": "Intermediate",
        "Last Suitability Test": "2026-01-22",
        "Suitability Outcome": "Aligned with Balanced mandate",
    },
    "CH-priv-0512": {  # Purushottam R Sharma — Balanced, R-SUIT-ALLOC breach (needs remediation)
        # Breach numbers moved here verbatim from the old Fig1..3 columns.
        "Mandate Type": "Discretionary — Balanced Growth",
        "Mand1 Label": "Equity",
        "Mand1 Value": "72% (ceiling 60%)",
        "Mand1 Status": "danger",
        "Mand1 Pct": 72,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "20% (band ≤35%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 20,
        "Mand3 Label": "Cash",
        "Mand3 Value": "8% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 8,
        # Perf1..3 already populated for this row — left untouched below.
        "Client Categorization": "Retail — MiFID standard",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — Single-name concentration (TechCorp)",
        "Hold1 Value": "22%",
        "Hold1 Status": "warn",
        "Hold1 Pct": 22,
        "Hold2 Label": "Fixed Income — EM Sovereign",
        "Hold2 Value": "20%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 20,
        "Hold3 Label": "Alternatives — Hedge Funds",
        "Hold3 Value": "15%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 15,
        "Risk Tolerance": "Balanced",
        "Investment Horizon": "5–10 years",
        "Knowledge Experience": "Intermediate",
        "Last Suitability Test": "2025-09-10",
        "Suitability Outcome": "Review needed — allocation drift from mandate",
    },
    "CH-priv-0658": {  # Aaron Lim Wei Jian — Growth, R-REG-NONDOM (needs remediation)
        # Categorization mismatch moved here verbatim from the old Fig1..3 columns.
        "Mandate Type": "Discretionary — Growth",
        "Mand1 Label": "Equity",
        "Mand1 Value": "68% (ceiling 75%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 68,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "20% (band ≤30%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 20,
        "Mand3 Label": "Cash",
        "Mand3 Value": "12% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 12,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+8.9%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — MSCI ACWI",
        "Perf2 Value": "+8.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "+0.5 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Retail",
        "Category Required": "Elective-professional",
        "Category Review Status": "Rule KB-REG-2026-04 v2 (FCA) — reassessment required",
        "Hold1 Label": "Equities — Non-dom offshore structures",
        "Hold1 Value": "30%",
        "Hold1 Status": "warn",
        "Hold1 Pct": 30,
        "Hold2 Label": "Fixed Income — Global Aggregate",
        "Hold2 Value": "25%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 25,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "12%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 12,
        "Risk Tolerance": "Growth",
        "Investment Horizon": "10+ years",
        "Knowledge Experience": "Advanced",
        "Last Suitability Test": "2025-12-01",
        "Suitability Outcome": "Aligned with Growth mandate",
    },
    "CH-priv-0847": {  # Sophie Hofer — Conservative, doc-expiry (needs remediation)
        "Mandate Type": "Advisory — Conservative",
        "Mand1 Label": "Equity",
        "Mand1 Value": "28% (ceiling 40%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 28,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "58% (band ≤65%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 58,
        "Mand3 Label": "Cash",
        "Mand3 Value": "14% (band ≥10%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 14,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+3.1%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 20/80 blended",
        "Perf2 Value": "+3.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "-0.3 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Retail — MiFID standard",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Fixed Income — CHF Government",
        "Hold1 Value": "40%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 40,
        "Hold2 Label": "Fixed Income — IG Corporate",
        "Hold2 Value": "22%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 22,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "14%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 14,
        "Risk Tolerance": "Conservative",
        "Investment Horizon": "3–5 years",
        "Knowledge Experience": "Basic",
        "Last Suitability Test": "2026-02-18",
        "Suitability Outcome": "Aligned with Conservative mandate",
    },
    "CH-priv-0093": {  # Markus Brunner — Balanced, R-SUIT-REVIEW (needs remediation)
        # Snapshot moved here verbatim from the old Fig1..3 columns.
        "Mandate Type": "Discretionary — Balanced",
        "Mand1 Label": "Equity",
        "Mand1 Value": "50% (ceiling 60%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 50,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "36% (band ≤45%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 36,
        "Mand3 Label": "Cash",
        "Mand3 Value": "14% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 14,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+5.9%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 60/40 blended",
        "Perf2 Value": "+6.1%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "-0.2 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Retail — MiFID standard",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — European Large Cap",
        "Hold1 Value": "28%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 28,
        "Hold2 Label": "Fixed Income — EUR Corporate",
        "Hold2 Value": "26%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 26,
        "Hold3 Label": "Multi-Asset Funds",
        "Hold3 Value": "16%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 16,
        "Risk Tolerance": "Balanced",
        "Investment Horizon": "5–10 years",
        "Knowledge Experience": "Intermediate",
        "Last Suitability Test": "2026-07-20",
        "Suitability Outcome": "Annual review due — ongoing, needs-based cadence",
    },
    "CH-priv-0774": {  # Victor Petrov — Growth, sow-refresh (needs remediation)
        "Mandate Type": "Discretionary — Growth",
        "Mand1 Label": "Equity",
        "Mand1 Value": "70% (ceiling 75%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 70,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "20% (band ≤30%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 20,
        "Mand3 Label": "Cash",
        "Mand3 Value": "10% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 10,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+10.1%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — MSCI ACWI",
        "Perf2 Value": "+8.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "+1.7 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Professional client (per-se)",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — Global Large Cap",
        "Hold1 Value": "40%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 40,
        "Hold2 Label": "Private Equity",
        "Hold2 Value": "22%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 22,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "10%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 10,
        "Risk Tolerance": "Growth",
        "Investment Horizon": "10+ years",
        "Knowledge Experience": "Advanced",
        "Last Suitability Test": "2025-10-05",
        "Suitability Outcome": "Aligned with Growth mandate",
    },
    "CH-corp-0901": {  # Rossi Capital S.p.A. — Balanced, compliant
        "Mandate Type": "Advisory — Balanced",
        "Mand1 Label": "Equity",
        "Mand1 Value": "52% (ceiling 60%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 52,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "34% (band ≤45%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 34,
        "Mand3 Label": "Cash",
        "Mand3 Value": "14% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 14,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+6.0%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 60/40 blended",
        "Perf2 Value": "+6.1%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "-0.1 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Professional client (per-se)",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — Eurozone Large Cap",
        "Hold1 Value": "30%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 30,
        "Hold2 Label": "Fixed Income — EUR Government",
        "Hold2 Value": "28%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 28,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "14%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 14,
        "Risk Tolerance": "Balanced",
        "Investment Horizon": "5–10 years",
        "Knowledge Experience": "Advanced — corporate treasury",
        "Last Suitability Test": "2026-05-12",
        "Suitability Outcome": "Aligned with Balanced mandate",
    },
    "CH-corp-0902": {  # Yamamoto K.K. — Conservative, compliant
        "Mandate Type": "Advisory — Conservative",
        "Mand1 Label": "Equity",
        "Mand1 Value": "25% (ceiling 40%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 25,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "60% (band ≤65%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 60,
        "Mand3 Label": "Cash",
        "Mand3 Value": "15% (band ≥10%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 15,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+2.8%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 20/80 blended",
        "Perf2 Value": "+3.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "-0.6 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Professional client (per-se)",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Fixed Income — JGB",
        "Hold1 Value": "42%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 42,
        "Hold2 Label": "Fixed Income — Global Aggregate",
        "Hold2 Value": "20%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 20,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "15%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 15,
        "Risk Tolerance": "Conservative",
        "Investment Horizon": "3–5 years",
        "Knowledge Experience": "Intermediate — corporate treasury",
        "Last Suitability Test": "2026-04-08",
        "Suitability Outcome": "Aligned with Conservative mandate",
    },
    "CH-corp-0903": {  # Thameside Ventures Ltd — Growth, compliant
        "Mandate Type": "Discretionary — Growth",
        "Mand1 Label": "Equity",
        "Mand1 Value": "72% (ceiling 75%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 72,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "18% (band ≤30%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 18,
        "Mand3 Label": "Cash",
        "Mand3 Value": "10% (band ≥5%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 10,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+11.4%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — MSCI ACWI",
        "Perf2 Value": "+8.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "+3.0 pp",
        "Perf3 Status": "ok",
        "Perf3 Pct": None,
        "Client Categorization": "Professional client (per-se)",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Equities — Global Large Cap",
        "Hold1 Value": "44%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 44,
        "Hold2 Label": "Private Equity",
        "Hold2 Value": "20%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 20,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "10%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 10,
        "Risk Tolerance": "Growth",
        "Investment Horizon": "10+ years",
        "Knowledge Experience": "Advanced — corporate treasury",
        "Last Suitability Test": "2026-06-01",
        "Suitability Outcome": "Aligned with Growth mandate",
    },
    "CH-priv-0355": {  # Katharina Vogt — Conservative, compliant
        "Mandate Type": "Advisory — Conservative",
        "Mand1 Label": "Equity",
        "Mand1 Value": "22% (ceiling 40%)",
        "Mand1 Status": "ok",
        "Mand1 Pct": 22,
        "Mand2 Label": "Fixed income",
        "Mand2 Value": "62% (band ≤65%)",
        "Mand2 Status": "ok",
        "Mand2 Pct": 62,
        "Mand3 Label": "Cash",
        "Mand3 Value": "16% (band ≥10%)",
        "Mand3 Status": "ok",
        "Mand3 Pct": 16,
        "Perf1 Label": "Portfolio return (YTD)",
        "Perf1 Value": "+2.5%",
        "Perf1 Status": "ok",
        "Perf1 Pct": None,
        "Perf2 Label": "Benchmark — 20/80 blended",
        "Perf2 Value": "+3.4%",
        "Perf2 Status": "ok",
        "Perf2 Pct": None,
        "Perf3 Label": "Variance vs benchmark",
        "Perf3 Value": "-0.9 pp",
        "Perf3 Status": "warn",
        "Perf3 Pct": None,
        "Client Categorization": "Retail — MiFID standard",
        "Category Required": "—",
        "Category Review Status": "Confirmed",
        "Hold1 Label": "Fixed Income — CHF Government",
        "Hold1 Value": "45%",
        "Hold1 Status": "ok",
        "Hold1 Pct": 45,
        "Hold2 Label": "Fixed Income — IG Corporate",
        "Hold2 Value": "20%",
        "Hold2 Status": "ok",
        "Hold2 Pct": 20,
        "Hold3 Label": "Cash & Money Market",
        "Hold3 Value": "16%",
        "Hold3 Status": "ok",
        "Hold3 Pct": 16,
        "Risk Tolerance": "Conservative",
        "Investment Horizon": "3–5 years",
        "Knowledge Experience": "Basic",
        "Last Suitability Test": "2026-03-15",
        "Suitability Outcome": "Aligned with Conservative mandate",
    },
}

# Retire these cases' use of the case-specific Fig1..3 slot now that their
# content lives in a generic always-shown section instead. Clearing them
# out avoids a stale, orphaned figure rendering behind the new sections.
CLEAR_FIG_FOR_REFS = ["CH-priv-0512", "CH-priv-0658", "CH-priv-0093"]
FIG_COLUMNS = [
    "Fig1 Label",
    "Fig1 Value",
    "Fig1 Status",
    "Fig1 Pct",
    "Fig2 Label",
    "Fig2 Value",
    "Fig2 Status",
    "Fig2 Pct",
    "Fig3 Label",
    "Fig3 Value",
    "Fig3 Status",
    "Fig3 Pct",
]


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Clients"]

    existing_headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    header_style_source = ws.cell(row=HEADER_ROW, column=2)  # any populated header cell

    start_col = ws.max_column + 1
    new_cols_added = [h for h in NEW_COLUMNS if h not in existing_headers]
    if not new_cols_added:
        print("All target columns already present — nothing to add, only refreshing values.")
    for offset, header in enumerate(new_cols_added):
        col = start_col + offset
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        # openpyxl's stubs type Cell.font/.fill/... as a narrower Serialisable
        # than what StyleProxy (returned by copy.copy on a style descriptor)
        # actually satisfies at runtime — a known stub gap, not a real bug.
        cell.font = copy.copy(header_style_source.font)  # pyright: ignore[reportAttributeAccessIssue]
        cell.fill = copy.copy(header_style_source.fill)  # pyright: ignore[reportAttributeAccessIssue]
        cell.border = copy.copy(header_style_source.border)  # pyright: ignore[reportAttributeAccessIssue]
        cell.alignment = copy.copy(header_style_source.alignment)  # pyright: ignore[reportAttributeAccessIssue]

    # Re-read the full header map now that new columns exist.
    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    col_by_header = {str(h): i + 1 for i, h in enumerate(headers) if h}
    ref_col = col_by_header["Client Ref"]

    updated = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        client_ref = ws.cell(row=r, column=ref_col).value
        if not isinstance(client_ref, str) or client_ref not in ROWS:
            continue
        for field, value in ROWS[client_ref].items():
            ws.cell(row=r, column=col_by_header[field], value=value)  # pyright: ignore
        if client_ref in CLEAR_FIG_FOR_REFS:
            for field in FIG_COLUMNS:
                # ws.cell(..., value=None) is a silent no-op in openpyxl (it only
                # assigns when value is not None) — must set .value directly.
                ws.cell(row=r, column=col_by_header[field]).value = None
        updated += 1

    seen_refs = {str(ws.cell(row=r, column=ref_col).value) for r in range(FIRST_DATA_ROW, ws.max_row + 1)}
    missing = set(ROWS) - seen_refs
    if missing:
        raise SystemExit(f"Client refs in ROWS but not found in the sheet: {sorted(missing)}")

    wb.save(EXCEL_PATH)
    print(f"Added {len(new_cols_added)} new columns, updated {updated} client rows.")
    print(f"→ {EXCEL_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
