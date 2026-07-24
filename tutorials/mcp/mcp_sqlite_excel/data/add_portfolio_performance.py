#!/usr/bin/env python3
"""One-off data-prep: add the "Portfolio Performance vs Benchmark" figure.

The portfolio-breach use case (R-SUIT-ALLOC) gets a second, optional figure
alongside its "Portfolio and Mandate" allocation bars (see
dashboard-v005/src/cases.json -> figure2_title). Same generic 3-row shape as
data/add_case_figures.py, just under a "perf" prefix instead of "fig" so both
figures can live on the same row without colliding:

    perf{n}_label, perf{n}_value, perf{n}_status ("ok" | "warn" | "danger" | "-"),
    perf{n}_pct (0-100, unused here but kept for shape parity)

Usage:
    uv run python data/add_portfolio_performance.py
    (then call reset_from_excel, or delete data/portfolio.db, to reload it)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(__file__).resolve().parent / "account_review_dataset.xlsx"

# client_ref -> [(label, value, status, pct), ...] (3 rows, pct optional/None)
PERFORMANCE: dict[str, list[tuple[str, str, str, float | None]]] = {
    # R-SUIT-ALLOC — same client as the "Portfolio and Mandate" allocation figure
    "CH-priv-0512": [
        ("Portfolio return (YTD)", "+4.1%", "ok", None),
        ("Benchmark — 60/40 blended", "+7.8%", "-", None),
        ("Variance vs benchmark", "-3.7 pp", "danger", None),
    ],
}

NEW_HEADERS = [
    "Perf1 Label",
    "Perf1 Value",
    "Perf1 Status",
    "Perf1 Pct",
    "Perf2 Label",
    "Perf2 Value",
    "Perf2 Status",
    "Perf2 Pct",
    "Perf3 Label",
    "Perf3 Value",
    "Perf3 Status",
    "Perf3 Pct",
]


def main() -> None:
    wb = load_workbook(XLSX_PATH)
    clients = wb["Clients"]

    client_header = [c.value for c in clients[4]]
    client_ref_col = client_header.index("Client Ref") + 1  # 1-based
    header_row_idx = 4
    first_new_col = len(client_header) + 1

    for offset, heading in enumerate(NEW_HEADERS):
        clients.cell(row=header_row_idx, column=first_new_col + offset, value=heading)

    matched = 0
    for row_idx in range(header_row_idx + 1, clients.max_row + 1):
        ref = clients.cell(row=row_idx, column=client_ref_col).value
        rows = PERFORMANCE.get(str(ref)) if ref else None
        if rows is None:
            continue
        matched += 1
        col = first_new_col
        for label, value, status, pct in rows:
            clients.cell(row=row_idx, column=col, value=label)
            clients.cell(row=row_idx, column=col + 1, value=value)
            clients.cell(row=row_idx, column=col + 2, value=status)
            clients.cell(row=row_idx, column=col + 3, value=pct)
            col += 4

    wb.save(XLSX_PATH)
    colum_range = f"{first_new_col}-{first_new_col + len(NEW_HEADERS) - 1}"
    print(f"Added performance rows onto {matched} clients (columns {colum_range}).")


if __name__ == "__main__":
    main()
