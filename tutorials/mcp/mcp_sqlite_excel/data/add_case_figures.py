#!/usr/bin/env python3
"""One-off data-prep: add per-case "figure" rows onto the Clients sheet.

Each RM use case surfaces a different structured figure in the client
record (see the Product mockups): an allocation-vs-mandate bar chart for
portfolio breaches, a documents/KYC checklist for expiry, a rule-impact
panel for regulatory change, etc. Rather than one bespoke column set per
case, every client gets the same generic 3-row shape:

    fig{n}_label, fig{n}_value, fig{n}_status ("ok" | "warn" | "danger" | "-"),
    fig{n}_pct (0-100, only meaningful for the allocation bars)

for n in (1, 2, 3). One reusable "case-figure" component (see
dashboard-v005/src/templates/figure_case.html) then renders whichever rows
are live for that row's rule_code — content differs, shape does not.

Usage:
    uv run python data/add_case_figures.py
    (then rebuild data/portfolio.db from the workbook, see merge_smart_actions.py)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(__file__).resolve().parent / "account_review_dataset.xlsx"

# client_ref -> [(label, value, status, pct), ...] (3 rows, pct optional/None)
FIGURES: dict[str, list[tuple[str, str, str, float | None]]] = {
    # R-SCR-ADVMEDIA (escalated)
    "CH-priv-0187": [
        ("Source", "WorldCheck", "-", None),
        ("Match", "Adverse media — ongoing litigation", "warn", None),
        ("Status", "With Compliance — client comment pending", "warn", None),
    ],
    # R-SCR-PEP (escalated)
    "CH-priv-0204": [
        ("Source", "WorldCheck", "-", None),
        ("Match", "PEP status — EDD refresh", "warn", None),
        ("Status", "With Compliance — EDD in progress", "warn", None),
    ],
    # R-SCR-ADVMEDIA (needs remediation)
    "CH-priv-0231": [
        ("Source", "WorldCheck (overnight screen)", "-", None),
        ("Match", "New adverse-media hit — litigation article", "danger", None),
        ("Status", "Pending RM review — confirm true match", "danger", None),
    ],
    # R-SUIT-ALLOC
    "CH-priv-0512": [
        ("Equity", "72% (ceiling 60%)", "danger", 72),
        ("Fixed income", "20% (band \u226435%)", "ok", 20),
        ("Cash", "8% (band \u22655%)", "ok", 8),
    ],
    # R-REG-NONDOM
    "CH-priv-0658": [
        ("Rule reference", "KB-REG-2026-04 v2 (FCA)", "danger", None),
        ("Category required", "Elective-professional", "danger", None),
        ("Current categorisation", "Retail", "danger", None),
    ],
    # R-DOC-EXPIRY
    "CH-priv-0847": [
        ("Passport", "Expires 2026-08-03 (13 days)", "danger", None),
        ("Proof of address", "On file — valid", "ok", None),
        ("Periodic KYC refresh", "Completed 2024 \u00b7 next due 2027", "ok", None),
    ],
    # R-SUIT-REVIEW
    "CH-priv-0093": [
        ("Risk profile on file", "Balanced", "ok", None),
        ("Last reviewed", "2026-07-20", "ok", None),
        ("Review cadence", "Ongoing, needs-based", "warn", None),
    ],
    # R-SOW-REFRESH
    "CH-priv-0774": [
        ("Incoming transfer", "CHF 2,400,000 (18 Jul)", "warn", None),
        ("Narrative on file", "Does not cover this event", "danger", None),
        ("Suggested evidence", "Company sale agreement", "-", None),
    ],
}

NEW_HEADERS = [
    "Fig1 Label",
    "Fig1 Value",
    "Fig1 Status",
    "Fig1 Pct",
    "Fig2 Label",
    "Fig2 Value",
    "Fig2 Status",
    "Fig2 Pct",
    "Fig3 Label",
    "Fig3 Value",
    "Fig3 Status",
    "Fig3 Pct",
]


def main() -> None:
    wb = load_workbook(XLSX_PATH)
    clients = wb["Clients"]

    client_header = [c.value for c in clients[4]]
    client_ref_col = client_header.index("Client Ref") + 1  # 1-based
    header_row_idx = 4
    first_new_col = len(client_header) + 1

    for offset, heading in enumerate(NEW_HEADERS):
        clients.cell(row=header_row_idx, column=first_new_col + offset, value=heading)

    matched = 0
    for row_idx in range(header_row_idx + 1, clients.max_row + 1):
        ref = clients.cell(row=row_idx, column=client_ref_col).value
        rows = FIGURES.get(str(ref)) if ref else None
        if rows is None:
            continue
        matched += 1
        col = first_new_col
        for label, value, status, pct in rows:
            clients.cell(row=row_idx, column=col, value=label)
            clients.cell(row=row_idx, column=col + 1, value=value)
            clients.cell(row=row_idx, column=col + 2, value=status)
            clients.cell(row=row_idx, column=col + 3, value=pct)
            col += 4

    wb.save(XLSX_PATH)
    print(f"Added figure rows onto {matched} clients (columns {first_new_col}-{first_new_col + len(NEW_HEADERS) - 1}).")


if __name__ == "__main__":
    main()
