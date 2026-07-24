#!/usr/bin/env python3
"""One-off data-prep: merge the "Smart Actions" sheet onto "Clients" in-place.

The dashboard binds a single ``clients`` list live via MCP (``list_rows``) for
the attention rail, portfolio table and client detail pages. To show a live,
per-use-case smart action (see the RM Account-Remediation Dashboard use-case
doc) without a SQL join, we denormalize the "Smart Actions" sheet onto
"Clients" ahead of time, keyed by Client Ref. This keeps the generic
mcp_sqlite_excel MCP server schema-agnostic — no join/view logic needed there.

Adds 5 columns to Clients (blank when a client has no open smart action):
    Action Owner, Action Title, Action Explanation, Action Button, Action Button Target

Usage:
    uv run python data/merge_smart_actions.py
    uv run python -m mcp_sqlite_excel.db.excel_loader  # (not needed; see below)

After running this, rebuild data/portfolio.db from the updated workbook, e.g.:
    uv run python -c "
from pathlib import Path
from mcp_sqlite_excel.db.excel_loader import bootstrap_from_excel
bootstrap_from_excel(Path('data/account_review_dataset.xlsx'), Path('data/portfolio.db'), replace=True)
"
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(__file__).resolve().parent / "account_review_dataset.xlsx"

NEW_HEADERS = [
    "Action Owner",
    "Action Title",
    "Action Explanation",
    "Action Button",
    "Action Button Target",
]


def main() -> None:
    wb = load_workbook(XLSX_PATH)

    clients = wb["Clients"]
    actions = wb["Smart Actions"]

    # Smart Actions: row 1 title, row 2 blank, row 3 header, rows 4+ data.
    action_header = [c.value for c in actions[3]]
    ref_col = action_header.index("Client Ref")
    owner_col = action_header.index("Owner")
    title_col = action_header.index("Action Title")
    expl_col = action_header.index("Explanation")
    button_col = action_header.index("Button")
    target_col = action_header.index("Button Target")

    by_ref: dict[str, tuple] = {}
    for row in actions.iter_rows(min_row=4, values_only=True):
        if not row or row[ref_col] in (None, ""):
            continue
        by_ref[str(row[ref_col])] = (
            row[owner_col],
            row[title_col],
            row[expl_col],
            row[button_col],
            row[target_col],
        )

    # Clients: row 1 title, rows 2-3 blank/section, row 4 header, rows 5+ data.
    client_header = [c.value for c in clients[4]]
    client_ref_col = client_header.index("Client Ref") + 1  # 1-based for openpyxl
    header_row_idx = 4
    first_new_col = len(client_header) + 1  # 1-based, right after last existing column

    for offset, heading in enumerate(NEW_HEADERS):
        clients.cell(row=header_row_idx, column=first_new_col + offset, value=heading)

    matched = 0
    for row_idx in range(header_row_idx + 1, clients.max_row + 1):
        ref = clients.cell(row=row_idx, column=client_ref_col).value
        if ref in (None, ""):
            continue
        values = by_ref.get(str(ref))
        if values is None:
            continue
        matched += 1
        for offset, value in enumerate(values):
            clients.cell(row=row_idx, column=first_new_col + offset, value=value)

    wb.save(XLSX_PATH)
    print(
        f"Merged {matched} smart actions onto Clients (columns {first_new_col}-{first_new_col + len(NEW_HEADERS) - 1})."
    )


if __name__ == "__main__":
    main()
